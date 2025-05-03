#!/usr/bin/env python3
"""
Excel Sheet Scanner

This script scans all Excel files in a specified directory,
extracts their sheet information, and saves it to a text file.
"""

import os
import sys
from pathlib import Path
import openpyxl
import pandas as pd

def scan_excel_files(directory_path, output_file):
    """
    Scan all Excel files in the specified directory and save their sheet information to a text file.
    
    Args:
        directory_path (str): Path to the directory containing Excel files
        output_file (str): Path to the output text file
    """
    directory = Path(directory_path)
    
    # Check if directory exists
    if not directory.exists() or not directory.is_dir():
        print(f"Error: Directory {directory_path} does not exist or is not a directory")
        return
    
    # Find all Excel files in the directory
    excel_files = []
    for file_path in directory.iterdir():
        if file_path.is_file() and file_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
            excel_files.append(file_path)
    
    if not excel_files:
        print(f"No Excel files found in {directory_path}")
        return
    
    # Open the output file for writing
    with open(output_file, 'w') as f:
        f.write("Excel Files and Sheets Information\n")
        f.write("================================\n\n")
        
        # Process each Excel file
        for excel_file in excel_files:
            try:
                # Open the Excel file with openpyxl for detailed information
                workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                sheet_names = workbook.sheetnames
                
                # Write file information to the output file
                f.write(f"File: {excel_file.name}\n")
                f.write("Sheets:\n")
                
                # Write sheet information from openpyxl
                for sheet_name in sheet_names:
                    sheet = workbook[sheet_name]
                    
                    # Get the dimensions of the sheet
                    min_row = sheet.min_row
                    max_row = sheet.max_row
                    min_col = sheet.min_column
                    max_col = sheet.max_column
                    
                    # Write sheet details
                    f.write(f"  - {sheet_name} (Rows: {max_row-min_row+1}, Columns: {max_col-min_col+1})\n")
                    
                    # Try to get some sample data from the sheet
                    f.write(f"    Sample data (first 5 rows, first 5 columns):\n")
                    
                    # Get column headers (first row)
                    headers = []
                    for col in range(min_col, min(min_col + 5, max_col + 1)):
                        cell_value = sheet.cell(min_row, col).value
                        headers.append(str(cell_value) if cell_value is not None else "")
                    
                    if any(headers):  # Only write if there are any non-empty headers
                        f.write(f"    Headers: {', '.join(h for h in headers if h)}\n")
                
                # Close the openpyxl workbook
                workbook.close()
                
                # Also use pandas to get additional information about the Excel file
                try:
                    # Read the Excel file with pandas
                    excel_data = pd.ExcelFile(excel_file)
                    
                    # Get sheet names from pandas
                    pandas_sheet_names = excel_data.sheet_names
                    
                    f.write("\n  Additional sheet information (from pandas):\n")
                    
                    # Process each sheet
                    for sheet_name in pandas_sheet_names:
                        # Read the sheet into a DataFrame
                        df = pd.read_excel(excel_data, sheet_name=sheet_name)
                        
                        # Get information about the DataFrame
                        num_rows, num_cols = df.shape
                        
                        f.write(f"    - {sheet_name}: {num_rows} rows, {num_cols} columns\n")
                        
                        # Get column names
                        column_names = df.columns.tolist()
                        if column_names:
                            f.write(f"      Column names: {', '.join(str(col) for col in column_names[:5])}")
                            if len(column_names) > 5:
                                f.write(f" ... ({len(column_names) - 5} more)")
                            f.write("\n")
                        
                        # Check for NaN values
                        nan_count = df.isna().sum().sum()
                        if nan_count > 0:
                            f.write(f"      Contains {nan_count} empty cells\n")
                        
                        # Check data types
                        dtypes = df.dtypes.value_counts().to_dict()
                        if dtypes:
                            dtype_str = ", ".join(f"{count} {dtype}" for dtype, count in dtypes.items())
                            f.write(f"      Data types: {dtype_str}\n")
                    
                except Exception as e:
                    f.write(f"  Error getting pandas information: {str(e)}\n")
                
                f.write("\n")
                
            except Exception as e:
                f.write(f"Error processing {excel_file.name}: {str(e)}\n\n")
    
    print(f"Excel files information saved to {output_file}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python excel_sheet_scanner.py <directory_path> <output_file>")
        sys.exit(1)
    
    directory_path = sys.argv[1]
    output_file = sys.argv[2]
    
    scan_excel_files(directory_path, output_file)